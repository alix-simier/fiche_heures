from datetime import date, datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import calendar

JOURS_FR = [
    "lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"
]

MOIS_FR = [
    "janvier", "février", "mars", "avril", "mai", "juin",
    "juillet", "août", "septembre", "octobre", "novembre", "décembre"
]


def jours_du_mois(mois, annee):
    jours = []
    nb_jours = calendar.monthrange(annee, mois)[1]

    for jour in range(1, nb_jours + 1):
        d = date(annee, mois, jour)
        nom_jour = JOURS_FR[d.weekday()]
        nom_mois = MOIS_FR[mois - 1]
        jour_formate = f"{nom_jour} {jour} {nom_mois} {annee}"
        jours.append(jour_formate)

    return jours

def generation_template_feuille_rtt(ws, nom, prenom, mois, annee, societe):
    nom = nom.upper()
    prenom = prenom.upper()
    mois = mois
    annee = annee
    societe = societe
    date = datetime(annee, mois, 1)
    mois_annee = date.strftime('%b-%y').lower()
    nom_feuille = f"{nom} {prenom}"
    nom_feuille = nom_feuille[:31]
    ws.title = nom_feuille

    bold = Font(bold=True)
    bold_underlign = Font(bold=True, underline='single')
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thin_border_up_down = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws['A1'] = societe
    ws['A1'].font = bold
    ws.merge_cells('B1:C1')
    ws['B1'] = "NOM :"
    ws['B1'].font = bold
    ws['D1'] = nom
    ws['F1'] = mois_annee
    ws['F1'].font = bold_underlign

    ws.merge_cells('B2:C2')
    ws['B2'] = "PRENOM :"
    ws['B2'].font = bold
    ws['D2'] = prenom

    ws['F3'] = "NBRE D'HEURES"
    ws['F3'].font = bold
    ws['F3'].border = thin_border
    ws['G3'] = "H100%"
    ws['H3'] = "H125%"
    ws['I3'] = "RTT ACQUIS"
    ws['J3'] = "RTT PRIS"
    ws['K3'] = "PRIME"

    debut_semaine = 4
    date_ligne = {}
    l = 4
    jours = jours_du_mois(mois, annee)
    for idx, j in enumerate(jours):
        dernier = (idx == len(jours) - 1)
        current_date = datetime(int(annee), int(mois), int(idx+1))
        if not j.startswith("dimanche") and not dernier:
            date_ligne[current_date] = l
            ws[f'A{l}'] = j
            ws[f'A{l}'].border = thin_border
            ws[f'B{l}'].border = thin_border_up_down
            ws[f'C{l}'].border = thin_border_up_down
            ws[f'D{l}'].border = thin_border_up_down
            ws[f'E{l}'].border = thin_border_up_down
            ws[f'F{l}'].border = thin_border
            l += 1
        else :
            date_ligne[current_date] = l
            ws[f'A{l}'] = j
            ws[f'A{l}'].border = thin_border
            ws[f'B{l}'].border = thin_border_up_down
            ws[f'C{l}'].border = thin_border_up_down
            ws[f'D{l}'].border = thin_border_up_down
            ws[f'E{l}'].border = thin_border_up_down
            ws[f'F{l}'].border = thin_border
            l += 1
            ws[f'A{l}'].border = thin_border
            ws[f'B{l}'] = "TOTAL HEURES DE LA SEMAINE"
            ws[f'B{l}'].border = thin_border_up_down
            ws[f'C{l}'].border = thin_border_up_down
            ws[f'D{l}'].border = thin_border_up_down
            ws[f'E{l}'] = f'=SUM(F{debut_semaine}:F{l-1})' 
            ws[f'E{l}'].border = thin_border
            ws[f'F{l}'] = f"""=IF(I{l}>0," DONT " & I{l} & "H RTT ACQUIS","")"""
            ws[f'F{l}'].border = thin_border
            ws[f'G{l}'] = f'=IF(E{l}<=35,E{l},35)' 
            ws[f'H{l}'] = f'=IF(AND(39<=E{l},E{l}<=43),E{l}-39, IF(E{l}>43, 4, 0))'
            ws[f'I{l}'] = f'=IF(AND(35<E{l},E{l}<=39),E{l}-35, IF(E{l}<=35,0, 4))' 
            ws[f'K{l}'] = f'=IF(E{l}>43,E{l}-43,0)'
            l += 1
            debut_semaine = l
    l+=1

    ws[f'A{l}'] = "SOLDE RTT DU AU 01/10/25 :"
    ws[f'A{l}'].font = bold
    ws[f'B{l}'].font = bold
    ws[f'D{l}'] = "TOTAL HEURES A 100% :"
    ws[f'D{l}'].font = bold
    ws[f'F{l}'] = f"=G{l}"
    ws[f'F{l}'].font = bold
    ws[f'G{l}'] = f"=SUM(G4:G{l-2})"
    ws[f'G{l}'].font = bold
    ws[f'H{l}'] = f"=SUM(H4:H{l-2})"
    ws[f'H{l}'].font = bold
    ws[f'I{l}'] = f"=SUM(I4:I{l-2})"
    ws[f'I{l}'].font = bold
    ws[f'J{l}'] = f"=SUM(J4:J{l-2})"
    ws[f'J{l}'].font = bold
    ws[f'K{l}'] = f"=SUM(K4:K{l-2})"
    ws[f'K{l}'].font = bold
    l+=1

    ws[f'A{l}'] = "RTT ACQUIS EN OCT 25 : "
    ws[f'A{l}'].font = bold
    ws[f'B{l}'] = f"=I{l-1}"
    ws[f'D{l}'] = "TOTAL HEURES A 125% :"
    ws[f'D{l}'].font = bold
    ws[f'F{l}'] = f"= H{l-1}"
    ws[f'F{l}'].font = bold
    ws[f'G{l}'] = "H100%"
    ws[f'H{l}'] = "H125%"
    ws[f'I{l}'] = "RTT ACQUIS"
    ws[f'J{l}'] = "RTT PRIS"
    ws[f'K{l}'] = "PRIME"
    l += 1

    ws[f'A{l}'] = "RTT PRIS EN OCT 25 :"
    ws[f'A{l}'].font = bold
    ws[f'B{l}'] = f'=J{l-2}'
    ws[f'B{l}'].font = bold
    ws[f'D{l}'] = "TOTAL GENERAL :"
    ws[f'D{l}'].font = bold
    ws[f'F{l}'] = f"=F{l-1}+F{l-2}"
    ws[f'F{l}'].font = bold
    l += 1

    ws[f'A{l}'] = f"SOLDE RTT DU AU {jours[-1]}"
    ws[f'A{l}'].font = bold
    ws[f'B{l}'] = f'=B{l-3}+B{l-2}-B{l-1}'
    ws[f'B{l}'].font = bold
    ws[f'C{l}'] = "Payés à 125%"
    ws[f'C{l}'].font = bold
    l += 2

    ws.merge_cells(f'A{l}:F{l}')
    ws[f'A{l}'] = "A retourner en fin de mois dûment remplie et signée SVP."
    ws[f'A{l}'].font = bold_underlign
    ws[f'A{l}'].alignment = center_alignment
    l += 2

    ws[f"A{l}"] = "Signature du salarié"

    return ws, date_ligne


def generation_template_feuille_sans_rtt(ws, nom, prenom, mois, annee, societe):
    nom = nom.upper()
    prenom = prenom.upper()
    mois = mois
    annee = annee
    societe = societe
    date = datetime(annee, mois, 1)
    mois_annee = date.strftime('%b-%y').lower()
    nom_feuille = f"{nom} {prenom}"
    nom_feuille = nom_feuille[:31]
    ws.title = nom_feuille

    bold = Font(bold=True)
    bold_underlign = Font(bold=True, underline='single')
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thin_border_up_down = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws['A1'] = societe
    ws['A1'].font = bold
    ws.merge_cells('B1:C1')
    ws['B1'] = "NOM :"
    ws['B1'].font = bold
    ws['D1'] = nom
    ws['F1'] = mois_annee
    ws['F1'].font = bold_underlign

    ws.merge_cells('B2:C2')
    ws['B2'] = "PRENOM :"
    ws['B2'].font = bold
    ws['D2'] = prenom

    ws['F3'] = "NBRE D'HEURES"
    ws['F3'].font = bold
    ws['F3'].border = thin_border
    ws['G3'] = "H100%"
    ws['H3'] = "H125%"
    ws['I3'] = "PRIME"

    debut_semaine = 4
    date_ligne = {}
    l = 4
    jours = jours_du_mois(mois, annee)
    for idx, j in enumerate(jours):
        dernier = (idx == len(jours) - 1)
        current_date = datetime(int(annee), int(mois), int(idx+1))
        if not j.startswith("dimanche") and not dernier:
            date_ligne[current_date] = l
            ws[f'A{l}'] = j
            ws[f'A{l}'].border = thin_border
            ws[f'B{l}'].border = thin_border_up_down
            ws[f'C{l}'].border = thin_border_up_down
            ws[f'D{l}'].border = thin_border_up_down
            ws[f'E{l}'].border = thin_border_up_down
            ws[f'F{l}'].border = thin_border
            l += 1
        else :
            date_ligne[current_date] = l
            ws[f'A{l}'] = j
            ws[f'A{l}'].border = thin_border
            ws[f'B{l}'].border = thin_border_up_down
            ws[f'C{l}'].border = thin_border_up_down
            ws[f'D{l}'].border = thin_border_up_down
            ws[f'E{l}'].border = thin_border_up_down
            ws[f'F{l}'].border = thin_border
            l += 1
            ws[f'A{l}'].border = thin_border
            ws[f'B{l}'] = "TOTAL HEURES DE LA SEMAINE"
            ws[f'B{l}'].border = thin_border_up_down
            ws[f'C{l}'].border = thin_border_up_down
            ws[f'D{l}'].border = thin_border_up_down
            ws[f'E{l}'] = f'=SUM(F{debut_semaine}:F{l-1})' 
            ws[f'E{l}'].border = thin_border
            ws[f'F{l}'].border = thin_border 
            ws[f'G{l}'] = f'=IF(E{l}<=35,E{l},35)' 
            ws[f'H{l}'] = f'=IF(AND(35<=E{l},E{l}<=43),E{l}-35, IF(E{l}>43, 8, 0))' 
            ws[f'I{l}'] = f'=IF(E{l}>43,E{l}-43,0)'
            l += 1
            debut_semaine = l
    l+=1

    ws[f'D{l}'] = "TOTAL HEURES A 100% :"
    ws[f'D{l}'].font = bold
    ws[f'F{l}'] = f"=G{l}"
    ws[f'F{l}'].font = bold
    ws[f'G{l}'] = f"=SUM(G4:G{l-2})"
    ws[f'G{l}'].font = bold
    ws[f'H{l}'] = f"=SUM(H4:H{l-2})"
    ws[f'H{l}'].font = bold
    ws[f'I{l}'] = f"=SUM(I4:I{l-2})"
    ws[f'I{l}'].font = bold
    l+=1

    ws[f'D{l}'] = "TOTAL HEURES A 125% :"
    ws[f'D{l}'].font = bold
    ws[f'F{l}'] = f"= H{l-1}"
    ws[f'F{l}'].font = bold
    ws[f'G{l}'] = "H100%"
    ws[f'H{l}'] = "H125%"
    ws[f'I{l}'] = "PRIME"
    l += 1

    ws[f'D{l}'] = "TOTAL GENERAL :"
    ws[f'D{l}'].font = bold
    ws[f'F{l}'] = f"=F{l-1}+F{l-2}"
    ws[f'F{l}'].font = bold
    l += 2

    ws.merge_cells(f'A{l}:F{l}')
    ws[f'A{l}'] = "A retourner en fin de mois dûment remplie et signée SVP."
    ws[f'A{l}'].font = bold_underlign
    ws[f'A{l}'].alignment = center_alignment
    l += 2

    ws[f"A{l}"] = "Signature du salarié"

    return ws, date_ligne