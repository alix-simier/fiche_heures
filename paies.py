import pandas as pd
from openpyxl import Workbook
from template_fiche_paie import generation_template_feuille_sans_rtt, generation_template_feuille_rtt

def fiche_paie(compte_travaux, regime_societe):
    compte_de_travaux = pd.read_excel(compte_travaux)
    employes = pd.read_excel(regime_societe)
    liste = []

    compte_de_travaux['Date'] = pd.to_datetime(compte_de_travaux['Date'])
    compte_de_travaux['Mois'] = compte_de_travaux['Date'].dt.month
    compte_de_travaux['Année'] = compte_de_travaux['Date'].dt.year
    
    # Use mode safely
    mois = compte_de_travaux['Mois'].mode()[0]
    année = compte_de_travaux['Année'].mode()[0]

    heures_jour_personne = compte_de_travaux.groupby(['Nom', 'Prénom', 'Date'], as_index=False)['Heures'].sum()
    personnes = heures_jour_personne.groupby(['Nom', 'Prénom'])

    wb_rtt = Workbook()
    wb_sans_rtt = Workbook()

    for (nom, prenom), group in personnes:
        ligne = employes[
            (employes['Nom'].str.lower() == nom.lower()) &
            (employes['Prenom'].str.lower() == prenom.lower())
        ]
        
        if ligne.empty:
            liste.append(f"{nom} {prenom}")
            continue

        entreprise = ligne['Entreprise'].values[0]
        regime = ligne['regime'].values[0]
        nom_feuille = f"{nom}_{prenom}"[:31]

        # LOGIC FIX: Create sheet for both cases
        if regime == "rtt":
            ws = wb_rtt.create_sheet(title=nom_feuille)
            ws, date_line = generation_template_feuille_rtt(ws, nom, prenom, mois, année, entreprise)
        elif regime == "sans rtt":
            ws = wb_sans_rtt.create_sheet(title=nom_feuille)
            ws, date_line = generation_template_feuille_sans_rtt(ws, nom, prenom, mois, année, entreprise)
        else:
            # On affiche le régime problématique
            print(f"⚠️ Erreur : Régime '{regime}' inconnu pour {prenom} {nom}. Pas de feuille créée.")
            date_line = None
            # Si tu es dans Streamlit, utilise plutôt :
            # st.warning(f"Régime '{regime}' non reconnu pour {prenom} {nom}.")

        # 3. On n'exécute la boucle que si la feuille a été générée avec succès
        if date_line is not None:
            for idx, ligne_group in group.iterrows():
                jour = ligne_group['Date']
                heures = ligne_group['Heures']
                if jour in date_line:
                    ligne_excel = date_line[jour]
                    ws[f'F{ligne_excel}'] = heures

    # Clean up default sheets
    for wb in [wb_rtt, wb_sans_rtt]:
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        if "TEMP" in wb.sheetnames:
            wb.remove(wb["TEMP"])

    # SAVE BEFORE RETURN
    wb_rtt.save("PAIE_RTT_2026.xlsx")
    wb_sans_rtt.save("PAIE_SANS_RTT_2026.xlsx")
    print("Fichiers Excel générés avec succès !")

    return wb_rtt, wb_sans_rtt, liste
    # Sauvegarde des deux fichiers séparés
    wb_rtt.save("01 - PAIE JANVIER - VALIDATION DES HEURES AVEC RTT 2026.xlsx")
    wb_sans_rtt.save("01 - PAIE JANVIER - VALIDATION DES HEURES SANS RTT 2026.xlsx")

    print("Fichiers Excel générés avec succès !")


if __name__ =="__main__":
    
    res = fiche_paie("ExportCTJANV2026.xlsx", "regime_societe.xlsx")
